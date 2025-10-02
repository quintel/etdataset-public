"""
Helper functions for ETLocal key calculations and template management.
"""

from typing import Any, Dict, List, Union, Optional
import pandas as pd
import yaml
from pathlib import Path
from src.yaml_calculator import YamlCalculator
import numpy as np
import warnings


def generate_etlocal_key(
    key,
    df_input_vars,
    yaml_folder=Path("config", "yaml_files_for_etlocal_key_calculation"),
):
    """
    Generate ETLocal key values using YamlCalculator from dependency YAML files.

    Parameters:
    -----------
    key : str
        The ETLocal key identifier
    df_input_vars : pandas.DataFrame
        DataFrame containing all input variables needed for calculations
    yaml_folder : pathlib.Path, default Path("config","yaml_files_for_etlocal_key_calculation")
        Path to the directory containing YAML dependency files

    Returns:
    --------
    pandas.DataFrame
        DataFrame containing calculated values for the ETLocal key

    Raises:
    -------
    FileNotFoundError
        If the YAML dependency file for the specified key doesn't exist
    """
    yaml_path = Path(yaml_folder, f"{key}_dependency.yaml")
    if not yaml_path.exists():
        raise FileNotFoundError(f"YAML file for key '{key}' not found at {yaml_path}")

    with open(yaml_path, "r") as file:
        yaml_defs = yaml.safe_load(file)

    calc = YamlCalculator(yaml_defs, debug=True)
    result_df = calc.calculate(df_input_vars, key)
    return result_df


def fill_template_with_etlocal_data(df_template_local, df_pipeline, etlocal_key=None):
    """
    Fill the ETLocal template with calculated values and commit messages.

    Parameters:
    -----------
    template_local : pandas.DataFrame
        The template dataframe with MultiIndex ['geo_id', 'group', 'subgroup', 'key']
    df_pipeline : pandas.DataFrame
        DataFrame containing the calculated values and commit messages
        Must have columns: 'etlocal_key', 'output_variable_value', 'commit'
    etlocal_key : str, optional
        If provided, returns only rows for this specific key

    Returns:
    --------
    pandas.DataFrame
        Updated template with filled values and commit messages
    """
    df_template_local_reset = df_template_local.reset_index()

    # Fill resetted template by looping over pipeline rows
    for code, row in df_pipeline.iterrows():
        value = row[etlocal_key]
        # value = row['output_variable_value']
        commit = row["commit"]

        # Select the relevant row in the template using a mask and update the value and commit
        mask = (df_template_local_reset["geo_id"] == code) & (
            df_template_local_reset["key"] == etlocal_key
        )
        df_template_local_reset.loc[mask, "value"] = value
        df_template_local_reset.loc[mask, "commit"] = commit

    # Reset MultiIndex including sort
    updated_template = df_template_local_reset.set_index(
        ["geo_id", "group", "subgroup", "key"]
    ).sort_index()

    return updated_template


def validate_template_data(template_data, etlocal_key, min_value=None, max_value=None):
    """
    Validate template data quality by checking various metrics.

    Parameters:
    -----------
    template_data : pandas.DataFrame
        Template data to validate.
    etlocal_key : str
        The ETLocal key being validated.
    min_value : float, optional
        Minimum expected value.
    max_value : float, optional
        Maximum expected value

    Returns:
    --------
    dict
        Validation results with basic statistics and value checks.
    """
    try:
        key_data = template_data.xs(etlocal_key, level="key")["value"]
    except KeyError:
        return {"error": f'Key "{etlocal_key}" not found in template'}

    # Convert values to native Python types for readability
    def to_native(value):
        return (
            None
            if pd.isna(value)
            else int(value) if isinstance(value, (np.integer, int)) else float(value)
        )

    results = {
        "key": etlocal_key,
        "total_rows": int(len(key_data)),
        "zero_count": int((key_data == 0).sum()),
        "nan_count": int(key_data.isna().sum()),
        "min_value": to_native(key_data.min()),
        "max_value": to_native(key_data.max()),
        "mean_value": to_native(key_data.mean()),
        "median_value": to_native(key_data.median()),
    }

    if min_value is not None:
        results["below_min"] = int((key_data < min_value).sum())
    if max_value is not None:
        results["above_max"] = int((key_data > max_value).sum())

    return results


def fill_template_with_uniform_value_etlocal_data(key, value, df_template_local):
    """
    Fill the ETLocal template with a uniform value for all municipalities for a specific key.

    This function is useful for setting values that are the same across all municipalities,
    such as setting coal demand to zero or applying a uniform policy value.

    Parameters:
    -----------
    key : str
        The ETLocal key identifier to fill in the template
    value : float or int
        The uniform value to assign to all municipalities for this key
    df_template_local : pandas.DataFrame
        The template dataframe with MultiIndex ['geo_id', 'group', 'subgroup', 'key']

    Returns:
    --------
    pandas.DataFrame
        Updated template with the uniform value filled for the specified key

    Raises:
    -------
    KeyError
        If the specified key is not found in the template

    Example:
    --------
    >>> # Set coal demand to zero for all municipalities
    >>> updated_template = fill_template_with_uniform_value_etlocal_data(
    ...     'buildings_final_demand_coal_demand',
    ...     0,
    ...     df_template_local
    ... )

    Notes:
    ------
    - This function modifies the template by setting the same value for all geo_ids
    - The commit message is automatically generated to document the uniform assignment
    - Useful for scenarios where certain energy types are not used in a region
    """
    # Create a copy to avoid modifying the original
    df_template_updated = df_template_local.copy()

    # Reset index for easier manipulation
    df_template_reset = df_template_updated.reset_index()

    # Create a mask to find all rows with the specified key
    mask = df_template_reset["key"] == key

    # Check if the key exists in the template
    if not mask.any():
        raise KeyError(f'Key "{key}" not found in template')

    # Fill the value for all municipalities with this key
    df_template_reset.loc[mask, "value"] = value

    # Generate a commit message documenting the uniform assignment
    if {value} == {60000}:  # specific commit for interconnector capacity
        commit_message = f"No data available. Set to a large value ({value} MW) for all municipalities to enable all electricity demand and production to be imported and exported."
    else:
        commit_message = f"No data available. Set to {value} for all municipalities."
    df_template_reset.loc[mask, "commit"] = commit_message

    # Reset MultiIndex and sort
    updated_template = df_template_reset.set_index(
        ["geo_id", "group", "subgroup", "key"]
    ).sort_index()

    return updated_template


def compare_template_with_verification(
    df_template_local,
    key,
    verification_file_path=Path("Specific for 2023 dataset update", "Values_2019_verification.xlsx"),
):
    """
    Compare ETLocal template values with verification data for a specific key across municipalities.

    This function extracts values for a specific ETLocal key from the template and compares
    them with corresponding values from a verification file, providing detailed analysis
    of differences between the two datasets. Missing municipalities are skipped.

    Parameters:
    -----------
    key : str
        The ETLocal key identifier to analyze
    df_template_local : pandas.DataFrame
        The template dataframe with MultiIndex ['geo_id', 'group', 'subgroup', 'key']
    verification_file_path : pathlib.Path, default Path("data", "reporting", "Values_2019_verification.xlsx")
        Path to the verification Excel file containing reference values

    Returns:
    --------
    pandas.DataFrame
        Comparison results with columns:
        - geo_id: Municipality identifier
        - template_value: Value from the template
        - verification_value: Value from verification file
        - absolute_diff: Absolute difference
        - relative_diff: Relative difference (as percentage)
        - status: Classification of the difference (e.g., 'match', 'small_diff', 'large_diff')
    """
    # Check if verification file exists
    if not verification_file_path.exists():
        raise FileNotFoundError(
            f"Verification file not found at {verification_file_path}"
        )

    # Extract template values for the specified key from multi-index
    try:
        template_key_data = df_template_local.xs(key, level="key")
        template_values = template_key_data["value"].copy()
    except KeyError:
        available_keys = df_template_local.index.get_level_values("key").unique()
        raise KeyError(
            f'Key "{key}" not found in template. Available keys: {sorted(available_keys)}'
        )

    # If template_values still has a MultiIndex, extract just the geo_id level
    if isinstance(template_values.index, pd.MultiIndex):
        template_values = template_values.reset_index()
        template_values = template_values.set_index("geo_id")["value"]

    # Load verification data
    try:
        if verification_file_path.suffix == ".xlsx":
            verification_df = pd.read_excel(verification_file_path)
        else:
            verification_df = pd.read_csv(verification_file_path)
    except Exception as e:
        raise FileNotFoundError(f"Error loading verification file: {e}")

    # Check if the key exists in verification data
    if key not in verification_df.columns:
        print(f"  Key '{key}' not found in verification data")
        return pd.DataFrame()  # Return empty DataFrame

    # Set up verification data with proper index
    geo_id_col = verification_df.columns[0]  # Assume first column is geo_id
    verification_df_indexed = verification_df.set_index(geo_id_col)
    verification_values = verification_df_indexed[key].copy()

    # Convert indices to same type (typically int for municipality codes)
    try:
        template_values = pd.Series(
            template_values.values, index=template_values.index.astype(int)
        )
        verification_values = pd.Series(
            verification_values.values, index=verification_values.index.astype(int)
        )
    except (ValueError, TypeError):
        try:
            template_values = pd.Series(
                template_values.values, index=template_values.index.astype(str)
            )
            verification_values = pd.Series(
                verification_values.values, index=verification_values.index.astype(str)
            )
        except Exception:
            return pd.DataFrame()

    # Find common municipalities and differences
    template_municipalities = set(template_values.index)
    verification_municipalities = set(verification_values.index)
    common_municipalities = template_municipalities.intersection(
        verification_municipalities
    )

    missing_in_verification = template_municipalities - verification_municipalities
    missing_in_template = verification_municipalities - template_municipalities

    # Print municipality analysis
    print(f"\nMUNICIPALITY COMPARISON FOR KEY: {key}")
    print(f"  Template municipalities: {len(template_municipalities)}")
    print(f"  Verification municipalities: {len(verification_municipalities)}")
    print(f"  Common municipalities: {len(common_municipalities)}")
    print(f"  Missing in verification: {len(missing_in_verification)}")
    print(f"  Missing in template: {len(missing_in_template)}")

    if missing_in_verification:
        missing_sorted = sorted(list(missing_in_verification))
        print(f"\n  MUNICIPALITIES IN TEMPLATE BUT NOT IN VERIFICATION:")
        # Show first 20, then indicate if there are more
        for i, mun in enumerate(missing_sorted[:20]):
            print(f"     {mun}")
        if len(missing_sorted) > 20:
            print(f"     ... and {len(missing_sorted) - 20} more municipalities")

    if missing_in_template:
        missing_sorted = sorted(list(missing_in_template))
        print(f"\n  MUNICIPALITIES IN VERIFICATION BUT NOT IN TEMPLATE:")
        # Show first 20, then indicate if there are more
        for i, mun in enumerate(missing_sorted[:20]):
            print(f"     {mun}")
        if len(missing_sorted) > 20:
            print(f"     ... and {len(missing_sorted) - 20} more municipalities")

    if len(common_municipalities) == 0:
        print("No common municipalities found - cannot perform comparison")
        return pd.DataFrame()

    # Filter to only common municipalities and align the data
    common_municipalities_sorted = sorted(common_municipalities)
    template_values_filtered = template_values.loc[common_municipalities_sorted]
    verification_values_filtered = verification_values.loc[common_municipalities_sorted]

    # Create comparison DataFrame
    comparison = pd.DataFrame(
        {
            "template_value": template_values_filtered,
            "verification_value": verification_values_filtered,
        },
        index=common_municipalities_sorted,
    )

    # Handle any remaining NaN values
    comparison = comparison.dropna()

    if len(comparison) == 0:
        print("No valid data pairs found after removing NaN values")
        return pd.DataFrame()

    # Calculate differences
    comparison["absolute_diff"] = abs(
        comparison["template_value"] - comparison["verification_value"]
    )

    # Calculate relative difference (handle division by zero)
    comparison["relative_diff"] = (
        (comparison["template_value"] - comparison["verification_value"])
        / comparison["verification_value"].replace(0, float("inf"))
    ) * 100

    # Replace infinite values with a large number for classification
    comparison["relative_diff"] = comparison["relative_diff"].replace(
        [float("inf"), -float("inf")], 999
    )

    # Add status classification
    def classify_difference(row):
        abs_diff = row["absolute_diff"]
        rel_diff = abs(row["relative_diff"])

        if abs_diff == 0:
            return "exact_match"
        elif rel_diff <= 1:  # Within 1%
            return "small_diff"
        elif rel_diff <= 10:  # Within 10%
            return "medium_diff"
        else:
            return "large_diff"

    comparison["status"] = comparison.apply(classify_difference, axis=1)

    # Show summary statistics
    status_counts = comparison["status"].value_counts()
    print(f"\n COMPARISON SUMMARY ({len(comparison)} municipalities):")
    for status, count in status_counts.items():
        percentage = (count / len(comparison)) * 100
        print(f"  {status}: {count} municipalities ({percentage:.1f}%)")

    # Statistical summary
    print(f"\n VALUE STATISTICS:")
    print(
        f"  Template:     Min={comparison['template_value'].min():.2f}, Max={comparison['template_value'].max():.2f}, Mean={comparison['template_value'].mean():.2f}"
    )
    print(
        f"  Verification: Min={comparison['verification_value'].min():.2f}, Max={comparison['verification_value'].max():.2f}, Mean={comparison['verification_value'].mean():.2f}"
    )
    print(
        f"  Abs diff:     Min={comparison['absolute_diff'].min():.2f}, Max={comparison['absolute_diff'].max():.2f}, Mean={comparison['absolute_diff'].mean():.2f}"
    )

    # Sort by largest absolute differences
    comparison_sorted = comparison.sort_values("absolute_diff", ascending=False)

    # Show top differences with municipality codes
    print(f"\n TOP 10 LARGEST DIFFERENCES:")
    top_diffs = comparison_sorted.head(10)
    for geo_id, row in top_diffs.iterrows():
        rel_diff_str = (
            f"{row['relative_diff']:.1f}%" if row["relative_diff"] != 999 else "inf%"
        )
        print(
            f"  Municipality {geo_id}: Template={row['template_value']:.2f}, Verification={row['verification_value']:.2f}, Diff={row['absolute_diff']:.2f} ({rel_diff_str})"
        )

    return comparison_sorted


def get_most_recent_year_data(df_with_code):
    """
    Extract the most recent year's data for each municipality from a DataFrame
    with year columns, returning the most recent value for each service type.

    Parameters:
    -----------
    df_with_code : pandas.DataFrame
        DataFrame with municipalities as index and year columns

    Returns:
    --------
    pandas.DataFrame
        DataFrame with most recent year data for each service type per municipality
    """
    import re

    # Skip non-data columns
    skip_columns = ["Municipality"]
    data_columns = [col for col in df_with_code.columns if col not in skip_columns]

    # Extract years and service types from column names
    year_service_map = {}
    service_types = set()

    for col in data_columns:
        # Check if column name contains a year (4-digit number between 2000-2030)
        year_match = re.search(r"\b(20[0-3]\d)\b", str(col))
        if year_match:
            year = int(year_match.group(1))
            # Remove year from column name to get service type
            service_type = col.replace(year_match.group(1), "").strip()

            # Clean up service type name (remove extra spaces, special characters)
            service_type = re.sub(r"\s+", " ", service_type).strip()

            service_types.add(service_type)

            if service_type not in year_service_map:
                year_service_map[service_type] = {}
            year_service_map[service_type][year] = col

    # For each municipality, find the most recent value for each service type
    result_data = []

    for idx in df_with_code.index:
        municipality_data = df_with_code.loc[idx]

        # Handle case where loc returns a DataFrame (duplicate indices)
        if isinstance(municipality_data, pd.DataFrame):
            municipality_data = municipality_data.iloc[0]

        row_data = {
            "GemeenteCode": idx,
            "Municipality": municipality_data.get("Municipality", ""),
        }

        # For each service type, find the most recent non-null value
        for service_type in service_types:
            if service_type in year_service_map:
                # Get all years for this service type, sorted most recent first
                years_for_service = sorted(
                    year_service_map[service_type].keys(), reverse=True
                )

                # Find the most recent non-null value
                most_recent_value = None
                most_recent_year = None
                source_column = None

                for year in years_for_service:
                    col_name = year_service_map[service_type][year]
                    value = municipality_data[col_name]

                    # Handle Series values (in case of duplicate indices)
                    if isinstance(value, pd.Series):
                        value = value.iloc[0] if not value.empty else None

                    # Check if value is valid
                    if pd.notna(value) and value != "":
                        most_recent_value = value
                        most_recent_year = year
                        source_column = col_name
                        break

                # Create clean column names for the service type
                clean_service_name = (
                    service_type.replace(" ", "_").replace("ë", "e").replace("ï", "i")
                )
                clean_service_name = re.sub(
                    r"[^\w]", "_", clean_service_name
                )  # Replace non-word chars with underscore
                clean_service_name = re.sub(r"_+", "_", clean_service_name).strip(
                    "_"
                )  # Remove multiple underscores

                # Add to row data
                row_data[f"{clean_service_name}_value"] = most_recent_value
                row_data[f"{clean_service_name}_year"] = most_recent_year
                row_data[f"{clean_service_name}_source"] = source_column

        result_data.append(row_data)

    # Convert to DataFrame
    df_result = pd.DataFrame(result_data)
    df_result.set_index("GemeenteCode", inplace=True)

    return df_result


def preprocess_missing_km_data(file_path, gemeenten_file_path=None):
    """
    Complete preprocessing pipeline for missing KM data files (like aardgas consumption).

    Parameters:
    -----------
    file_path : str or Path
        Path to the data Excel file
    gemeenten_file_path : str or Path, optional
        Path to the municipalities reference Excel file
        If None, defaults to "data/raw/Gemeenten alfabetisch 2023.xlsx"

    Returns:
    --------
    pandas.DataFrame
        Processed DataFrame with most recent year data
    """

    # Set default gemeenten file path if not provided
    if gemeenten_file_path is None:
        gemeenten_file_path = Path("data", "raw", "Gemeenten alfabetisch 2023.xlsx")

    # 1. Load the data
    df_data = pd.read_excel(file_path)
    df_data = df_data.replace("?", np.nan)

    # 2. Fix the header structure
    # Skip the first column and use modern pandas syntax
    # Create the first row by taking the first row and filling NaN values with values from the right
    first_row = df_data.iloc[0, 1:].bfill()  # Skip first column with iloc[0, 1:]
    df_data.iloc[0, 1:] = first_row

    # 3. Load municipality reference data
    df_gemeenten = pd.read_excel(gemeenten_file_path)

    # 4. Create merged column names from first two rows
    df_data_cols = (
        df_data.iloc[0].fillna("").astype(str)
        + " "
        + df_data.iloc[1].fillna("").astype(str)
    )
    df_data_cols = df_data_cols.str.strip()  # Remove extra whitespace

    # Set the first column name to 'Municipality'
    df_data_cols.iloc[0] = "Municipality"

    # 5. Create new dataframe with merged column names
    df_data_merged = df_data.iloc[2:].copy()
    df_data_merged.columns = df_data_cols
    df_data_merged.reset_index(drop=True, inplace=True)

    # 6. Merge with gemeente reference to get GemeenteCode
    df_data_with_code = df_data_merged.merge(
        df_gemeenten[["Gemeentenaam", "GemeentecodeGM"]],
        left_on="Municipality",
        right_on="Gemeentenaam",
        how="left",
    )

    # 7. Drop the duplicate municipality name column and reorder
    df_data_with_code = df_data_with_code.drop("Gemeentenaam", axis=1)
    df_data_with_code = df_data_with_code[
        ["GemeentecodeGM", "Municipality"]
        + [
            col
            for col in df_data_with_code.columns
            if col not in ["GemeentecodeGM", "Municipality"]
        ]
    ]

    # 8. Handle duplicate gemeente codes by keeping the first occurrence
    df_data_with_code = df_data_with_code.drop_duplicates(
        subset=["GemeentecodeGM"], keep="first"
    )

    # 9. Set GemeentecodeGM as index
    df_data_with_code.set_index("GemeentecodeGM", inplace=True)

    # 10. Extract most recent year data
    df_data_recent = get_most_recent_year_data(df_data_with_code)

    return df_data_recent


def map_km_variables_to_ivar(df_recent_data, conversion_file_path):
    """
    Map KM variables to internal variables using the conversion table

    Parameters:
    -----------
    df_recent_data : pd.DataFrame
        Processed data with KM variable columns (ending with _value, _year, _source)
    conversion_file_path : str or Path
        Path to the km_ivar_conversion.csv file

    Returns:
    --------
    pd.DataFrame
        Mapped dataframe with internal variable naming
    """

    # Load the conversion mapping
    df_conversion = pd.read_csv(conversion_file_path)

    # Create mapping dictionary: km-variable -> internal variable (ivar)
    km_to_ivar = dict(
        zip(df_conversion["km-variable"], df_conversion["internal variable (ivar)"])
    )

    # Create a new dataframe to store the mapped data
    df_mapped = df_recent_data.copy()

    # Get all columns that end with '_value' to identify the service types
    value_columns = [col for col in df_mapped.columns if col.endswith("_value")]

    # Extract service type names (remove '_value' suffix)
    service_types = [col.replace("_value", "") for col in value_columns]

    # Map each service type to its corresponding internal variable
    for service_type in service_types:
        # Check if this service type has a corresponding KM variable in the conversion table
        km_var_candidates = [
            km_var
            for km_var in km_to_ivar.keys()
            if service_type.lower().replace("_", "").replace(" ", "")
            in km_var.lower().replace("_", "").replace(" ", "")
        ]

        # If no direct match, try to find it by the service type name itself
        if not km_var_candidates:
            # Try direct lookup using service type as potential km variable
            if service_type in km_to_ivar:
                km_var_candidates = [service_type]

        # For aardgas specifically, map the service types we know
        service_to_km_mapping = {
            "commerciele_dienstverlening": "gascomdv",
            "publieke_dienstverlening": "gaspubldv",
        }

        service_key = service_type.lower().replace(" ", "_")
        if service_key in service_to_km_mapping:
            km_var_candidates = [service_to_km_mapping[service_key]]

        # Process the mapping if we found a match
        for km_var in km_var_candidates:
            if km_var in km_to_ivar:
                ivar_name = km_to_ivar[km_var]

                # Check if the service type columns exist in the data
                value_col = f"{service_type}_value"
                year_col = f"{service_type}_year"
                source_col = f"{service_type}_source"

                if value_col in df_mapped.columns:
                    # Rename the columns to match internal variable naming
                    df_mapped[f"{ivar_name}_value"] = df_mapped[value_col]
                    df_mapped[f"{ivar_name}_year"] = (
                        df_mapped[year_col] if year_col in df_mapped.columns else None
                    )
                    df_mapped[f"{ivar_name}_source"] = df_mapped.get(
                        source_col, "Fill-up data"
                    )

                    # Drop the original columns
                    cols_to_drop = [
                        col
                        for col in [value_col, year_col, source_col]
                        if col in df_mapped.columns
                    ]
                    df_mapped = df_mapped.drop(columns=cols_to_drop)

                    break  # Only use the first match

    return df_mapped


def merge_fillup_data_to_input_vars(df_input_vars, df_fillup_mapped, verbose=True):
    """
    Merge the mapped fillup data back into df_input_vars, overwriting existing columns

    Parameters:
    -----------
    df_input_vars : pd.DataFrame
        Main input variables dataframe
    df_fillup_mapped : pd.DataFrame
        Mapped fillup data with internal variable naming
    verbose : bool, optional
        Whether to print debug information (default: True)

    Returns:
    --------
    pd.DataFrame
        Updated input variables dataframe
    """

    # Make a copy to avoid modifying the original
    df_updated = df_input_vars.copy()

    # Get the relevant columns from the mapped data (only the internal variable columns)
    fillup_columns = [
        col
        for col in df_fillup_mapped.columns
        if col.endswith(("_value", "_year", "_source"))
    ]

    if verbose:
        print(f"Sample df_updated.index: {df_updated.index[:5].tolist()}")
        print(f"Sample df_fillup_mapped.index: {df_fillup_mapped.index[:5].tolist()}")

    # Track updates for validation
    updated_variables = {}

    # For each municipality in the fillup data, update the corresponding values
    for gemeente_code in df_fillup_mapped.index:
        if gemeente_code in df_updated.index:
            for col in fillup_columns:
                # Extract the base variable name (without _value, _year, _source suffix)
                base_var = col.rsplit("_", 1)[0]

                # Only update if we have a valid value
                if pd.notna(df_fillup_mapped.loc[gemeente_code, col]):
                    # If it's a value column, update the corresponding column in df_input_vars
                    if col.endswith("_value"):
                        if base_var in df_updated.columns:
                            df_updated.loc[gemeente_code, base_var] = (
                                df_fillup_mapped.loc[gemeente_code, col]
                            )
                            if verbose:
                                print(f"Updated {base_var} for {gemeente_code}")

                            # Track for validation
                            if base_var not in updated_variables:
                                updated_variables[base_var] = 0
                            updated_variables[base_var] += 1
        else:
            if verbose:
                print(f"Municipality {gemeente_code} not found in df_updated.index")

    # Validation summary
    if verbose:
        print("=== Validation: Updated municipalities ===")
        for var, count in updated_variables.items():
            print(f"{var}: {count} municipalities updated")

    return df_updated


def generate_commit_message(
    output_description,
    arithmetic,
    variables_description,
    sector,
    year,
    variables_source,
) -> str:
    """Generate commit message with flexible variable handling"""
    if len(variables_description) == 2:
        return f"{output_description} in the {sector} sector is calculated by {arithmetic}ing {variables_description[0]} ({variables_source[0]}, {year}) by the {variables_description[1]} ({variables_source[1]}, {year})."
    elif len(variables_description) == 3:
        return "Left to be written"
    elif len(variables_description) == 1:
        return f"{output_description} in the {sector} sector is based on {variables_description[0]} ({variables_source[0]}, {year})."
    else:
        return "Left to be written"


def generate_commit_message_v2(
    output_description, arithmic, vars_description, sector, year, vars_source
):
    def format_var(desc, source):
        return f"{desc} ({source}, {year})"

    # Handle None or empty input
    if not vars_description:
        return (
            f"{output_description} in the {sector} sector is directly retrieved from {vars_source[0]} ({year})."
            # f"({vars_source[0]}, {year})."
        )

    # 1 variable
    if len(vars_description) == 1:
        return (
            f"{output_description} in the {sector} sector is directly based on "
            f"{format_var(vars_description[0], vars_source[0])}."
        )

    # 2 variables
    elif len(vars_description) == 2:
        # Use arithmic[0] and arithmic[1] here assuming arithmic is a list with two elements
        if arithmic[-1] == "subtract":
            return (
                f"{output_description} in the {sector} sector is calculated by "
                f"{arithmic[1]}ing {format_var(vars_description[1], vars_source[1])} from "
                f"{format_var(vars_description[0], vars_source[0])}."
            )
        elif arithmic[-1] == "add":
            return (
                f"{output_description} in the {sector} sector is calculated by "
                f"{arithmic[1]}ing {format_var(vars_description[0], vars_source[0])} to "
                f"{format_var(vars_description[1], vars_source[1])}."
            )
        else:
            return (
                f"{output_description} in the {sector} sector is calculated by "
                f"{arithmic[1]}ing {format_var(vars_description[0], vars_source[0])} by "
                f"{format_var(vars_description[1], vars_source[1])}."
            )

    # 3 variables
    elif len(vars_description) == 3:
        return (
            f"{output_description} in the {sector} sector is calculated by first "
            f"{arithmic[0]}ing {format_var(vars_description[0], vars_source[0])} and "
            f"{format_var(vars_description[1], vars_source[1])}, and "
            f"{arithmic[2]}ing {format_var(vars_description[2], vars_source[2])}."
        )

    # 4 variables
    elif len(vars_description) == 4:
        return (
            f"{output_description} in the {sector} sector is calculated by {arithmic[0]}ing "
            f"{format_var(vars_description[0], vars_source[0])}, {format_var(vars_description[1], vars_source[1])}, "
            f"{format_var(vars_description[2], vars_source[2])}, and {format_var(vars_description[3], vars_source[3])}."
        )

    # 6 variables
    elif len(vars_description) == 6:
        return (
            f"{output_description} in the {sector} sector is calculated by {arithmic[0]}ing"
            f"{format_var(vars_description[0], vars_source[0])}, {format_var(vars_description[1], vars_source[1])}, "
            f"{format_var(vars_description[2], vars_source[2])}, {format_var(vars_description[3], vars_source[3])}, "
            f"{format_var(vars_description[4], vars_source[4])}, and "
            f"{format_var(vars_description[5], vars_source[5])}."
        )

    # Fallback
    else:
        return (
            f"{output_description} in the {sector} sector uses an unsupported number of variables "
            f"({len(vars_description)})."
        )


def generate_commits_yml(
    input_csv_paths: Union[List[Union[str, Path]], Union[str, Path]],
    output_dir: Union[str, Path],
    yml_filename: str = "commits.yml",
    analysis_year: Optional[int] = None,
) -> Dict[str, Any]:
    """
    Generate YAML file with commit messages and associated fields from CSV data.

    Args:
        input_csv_paths: Single path or list of paths to input CSV files
        output_dir: Directory to save output files
        yml_filename: Name for output YAML file
        analysis_year: Optional year to add to commits.yml with "Update to {analysis_year}" message

    Returns:
        Dictionary with results and file paths
    """
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    # Always use the output CSV from pivot_n_merge_processed
    if isinstance(input_csv_paths, (str, Path)):
        input_csv_paths = [input_csv_paths]

    input_path = Path(input_csv_paths[0])
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    try:
        df = pd.read_csv(input_path)
    except Exception as e:
        raise ValueError(f"Error reading CSV file: {e}")

    # Find the last row where geo_id == 'commit_message'
    commit_row = df[df["geo_id"] == "commit_message"]
    if commit_row.empty:
        raise ValueError("No commit_message row found in the provided CSV.")
    commit_row = commit_row.iloc[0]

    # Metadata columns to skip
    meta_cols = {"geo_id", "country", "name", "analysis_year"}

    # Group fields by commit message
    message_to_fields = {}
    for col in df.columns:
        if col in meta_cols:
            continue
        val = commit_row[col]
        if pd.notna(val) and str(val).strip() != "":
            msg = str(val)
            if msg not in message_to_fields:
                message_to_fields[msg] = []
            message_to_fields[msg].append(col)

    # Generate YAML content: group fields with the same message
    commits_file = "---\n"
    
    # Add analysis_year entry if provided
    if analysis_year is not None:
        commits_file += "- fields: \n"
        commits_file += f"  - analysis_year\n"
        commits_file += f'  message:\n   "Update to {analysis_year}"\n\n'
    
    for msg, fields in message_to_fields.items():
        commits_file += "- fields: \n"
        for key in fields:
            commits_file += f"  - {key}\n"
        commits_file += f'  message:\n   "{msg}"\n\n'

    # Save YAML file
    yml_output_path = output_path / yml_filename
    with open(yml_output_path, "w") as f:
        f.write(commits_file)

    # Remove the commit_message row and overwrite the CSV with the cleaned version
    df_clean = df[df["geo_id"] != "commit_message"].copy()
    df_clean.to_csv(input_path, index=False, na_rep="")

    return {"yml_path": str(yml_output_path)}


def pivot_n_merge_processed(
    input_csv_paths: Union[List[Union[str, Path]], Union[str, Path]],
    output_dir: Union[str, Path],
    country: str,
    analysis_year: int,
    config_path: Union[str, Path],
    csv_filename: str = "transformed_data.csv",
) -> Dict[str, Any]:
    """
    Most optimized version using reindex
    """

    if isinstance(input_csv_paths, (str, Path)):
        input_csv_paths = [input_csv_paths]

    input_paths = [Path(p) for p in input_csv_paths]

    # Collect ALL unique keys and geo_ids
    all_keys = set()
    all_geo_ids = set()
    all_data = []

    for input_path in input_paths:
        df = pd.read_csv(input_path)
        all_geo_ids.update(df["geo_id"].unique())

        df_with_keys = df[df["key"].notna()].copy()
        if len(df_with_keys) > 0:
            file_keys = set(df_with_keys["key"].unique())
            all_keys.update(file_keys)
            all_data.append(df_with_keys[["geo_id", "key", "value"]])

    all_keys.remove("analysis_year")
    print(f"Found {len(all_keys)} total unique keys")

    # Process data
    if all_data:
        # Combine all data
        combined_df = pd.concat(all_data, ignore_index=True)
        combined_df = (
            combined_df.groupby(["geo_id", "key"])["value"].first().reset_index()
        )

        # Pivot
        pivot_df = combined_df.pivot_table(
            index="geo_id", columns="key", values="value", aggfunc="first"
        ).reset_index()

        desired_columns = ["geo_id"] + sorted(list(all_keys))
        pivot_df = pivot_df.reindex(columns=desired_columns)

    else:
        desired_columns = ["geo_id"] + sorted(list(all_keys))
        pivot_df = pd.DataFrame(columns=desired_columns)
        pivot_df["geo_id"] = list(all_geo_ids)

    output_path = Path(output_dir)
    config_path = Path(config_path)
    output_path.mkdir(parents=True, exist_ok=True)

    try:
        config_df = pd.read_csv(config_path)
        geo_name_mapping = dict(zip(config_df["geo_id"], config_df["name"]))
    except Exception as e:
        raise ValueError(f"Error reading config file: {e}")

    base_df = pd.DataFrame(
        {
            "geo_id": sorted(list(all_geo_ids)),
            "country": country,
            "name": [
                geo_name_mapping.get(geo_id, geo_id)
                for geo_id in sorted(list(all_geo_ids))
            ],
            "analysis_year": analysis_year,
        }
    )

    result_df = base_df.merge(pivot_df, on="geo_id", how="left")
    result_df = result_df.sort_values("name").reset_index(drop=True)

    # --- Append commit messages row ---
    # Collect commit messages for each key from all input files
    commit_messages = {}
    for input_path in input_paths:
        df = pd.read_csv(input_path)
        for _, row in df.iterrows():
            key = row.get("key")
            commit = row.get("commit")
            if pd.notna(key) and key not in commit_messages and pd.notna(commit):
                commit_messages[key] = commit

    # Prepare the commit message row
    commit_row = {}
    for col in result_df.columns:
        if col == "geo_id":
            commit_row[col] = "commit_message"
        elif col == "country":
            commit_row[col] = ""
        elif col == "name":
            commit_row[col] = ""
        elif col == "analysis_year":
            commit_row[col] = ""
        else:
            commit_row[col] = commit_messages.get(col, "")

    # Append the row
    result_df = pd.concat([result_df, pd.DataFrame([commit_row])], ignore_index=True)

    csv_output_path = output_path / csv_filename
    result_df.to_csv(csv_output_path, index=False, na_rep="")

    final_data_columns = set(result_df.columns) - {
        "geo_id",
        "country",
        "name",
        "analysis_year",
    }
    print(f"Final result: {len(final_data_columns)} data columns")

    return {"csv_path": str(csv_output_path)}


def read_file(file_path: Union[str, Path]) -> pd.DataFrame:
    """
    Read either CSV or Excel file and return DataFrame.

    Args:
        file_path: Path to CSV or Excel file

    Returns:
        pd.DataFrame: Loaded data
    """
    import numpy as np
    import warnings

    file_path = Path(file_path)

    # Suppress openpyxl warnings temporarily
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
        if file_path.suffix.lower() in [".xlsx", ".xls"]:
            df = pd.read_excel(file_path)
        elif file_path.suffix.lower() == ".csv":
            df = pd.read_csv(file_path)
        else:
            raise ValueError(f"Unsupported file format: {file_path.suffix}")

    # Clean missing data markers immediately after loading
    missing_patterns = [
        "?",
        "nan",
        "NaN",
        "NA",
        "N/A",
        "not available",
        "Not Available",
        "",
        "-",
    ]
    for col in df.columns:
        for pattern in missing_patterns:
            mask = df[col].astype(str).str.strip() == pattern
            if mask.any():
                df.loc[mask, col] = np.nan

    # Ensure compatibility with openpyxl
    for col in df.columns:
        if df[col].dtype == "object":
            df[col] = df[col].where(pd.notna(df[col]), np.nan)

    return df


def create_column_mapping(
    main_cols: List[str], backup_cols: List[str], main_year: str = "2023"
) -> Dict[str, str]:
    """
    Create mapping between main file columns and backup file columns based on base names.

    Args:
        main_cols: Column names from main file
        backup_cols: Column names from backup file
        main_year: Year in main file columns to replace

    Returns:
        Dict mapping main_col -> backup_col
    """
    mapping = {}

    # Extract base names from main columns (remove year suffix)
    main_base_cols = {}
    for col in main_cols:
        if col.endswith(f"_{main_year}"):
            base_name = col.replace(f"_{main_year}", "")
            main_base_cols[base_name] = col

    # Find matching columns in backup
    for backup_col in backup_cols:
        # Try to find what base name this backup column represents
        for base_name, main_col in main_base_cols.items():
            # Check if backup column matches pattern: base_name_YEAR
            if backup_col.startswith(base_name + "_") and backup_col != main_col:
                # Extract year from backup column
                potential_year = backup_col.replace(base_name + "_", "")
                if potential_year.isdigit() and len(potential_year) == 4:
                    mapping[main_col] = backup_col
                    break

    return mapping


def generate_color_palette(num_sources):
    """
    Generate a color palette that scales with the number of sources.
    Uses HSV color space for even distribution of distinguishable colors.

    Args:
        num_sources (int): Number of different colors needed

    Returns:
        List[str]: List of hex color codes (without #)
    """
    import colorsys

    if num_sources == 1:
        return ["0000FF"]  # Standard blue for single source

    colors = []
    for i in range(num_sources):
        # Distribute hue evenly across color wheel (0-1)
        hue = i / num_sources
        # Fixed saturation and value for vibrant, readable colors
        saturation = 0.8
        value = 0.8

        # Convert HSV to RGB
        rgb = colorsys.hsv_to_rgb(hue, saturation, value)
        # Convert to hex format for Excel
        hex_color = "{:02X}{:02X}{:02X}".format(
            int(rgb[0] * 255), int(rgb[1] * 255), int(rgb[2] * 255)
        )
        colors.append(hex_color)

    return colors


def create_color_mapping(backup_paths):
    """
    Create color mapping based on backup file names.

    Args:
        backup_paths (List[Union[str, Path]]): List of backup file paths

    Returns:
        Dict[str, str]: Mapping from source name to hex color code
    """
    source_names = [Path(path).stem for path in backup_paths]
    colors = generate_color_palette(len(source_names))

    return dict(zip(source_names, colors))


def fill_missing_KM_data(
    main_csv_path: Union[str, Path],
    backup_csv_paths: List[Union[str, Path]],
    output_path: Union[str, Path],
    output_filename: str = "filled_data.xlsx",
    index_column: str = None,
    main_year: str = "2023",
    convert_remaining_to_zero: bool = True,
) -> Dict[str, Any]:
    """
    Fill missing values and create Excel file with bold formatting for filled data.
    Converts '?', '-', blank cells and similar markers to NaN at the start so they can be backfilled.
    """
    import numpy as np
    import warnings

    # Suppress pandas future warnings and openpyxl warnings
    warnings.filterwarnings("ignore", category=FutureWarning)
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    output_dir = Path(output_path)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Read and clean main file
    filled_df = read_file(main_csv_path)

    # Auto-detect index column if not specified
    if index_column is None:
        potential_id_cols = [
            col
            for col in filled_df.columns
            if any(
                pattern in col.lower()
                for pattern in ["id", "code", "key", "naam", "name"]
            )
        ]

        if potential_id_cols:
            index_column = potential_id_cols[0]
        else:
            index_column = filled_df.columns[0]

    if index_column not in filled_df.columns:
        raise ValueError(f"Index column '{index_column}' not found in main file.")

    # Count initial missing (after cleaning markers)
    initial_missing = filled_df.isnull().sum().sum()
    original_mask = filled_df.notnull()

    print(f"Initial missing: {initial_missing:,}")

    # Generate dynamic color mapping for sources
    color_map = create_color_mapping(backup_csv_paths)

    fills_by_source = {}
    filled_cells_by_source = {}

    # Fill from each backup
    for backup_path in backup_csv_paths:
        backup_path = Path(backup_path)

        try:
            backup_df = read_file(backup_path)  # This also cleans missing markers

            if index_column not in backup_df.columns:
                continue

            # Create column mapping
            column_mapping = create_column_mapping(
                filled_df.columns.tolist(), backup_df.columns.tolist(), main_year
            )

            if not column_mapping:
                continue

            # Track state before filling from this source
            before_mask = filled_df.notnull()
            source_name = backup_path.stem

            # Set up indices
            backup_df_indexed = backup_df.set_index(index_column)
            filled_df_indexed = filled_df.set_index(index_column)
            before_missing = filled_df.isnull().sum().sum()

            # Perform filling using column mapping
            for main_col, backup_col in column_mapping.items():
                if backup_col in backup_df_indexed.columns:
                    # Find rows where main is missing but backup has data
                    main_missing = filled_df_indexed[main_col].isnull()
                    backup_has_data = backup_df_indexed[backup_col].notnull()

                    fill_mask = main_missing & backup_has_data
                    fill_indices = fill_mask[fill_mask].index

                    if len(fill_indices) > 0:
                        # Ensure data type compatibility before assignment
                        backup_values = backup_df_indexed.loc[fill_indices, backup_col]

                        # Convert backup values to match main column type if possible
                        try:
                            if filled_df_indexed[main_col].dtype in [
                                "float64",
                                "int64",
                            ]:
                                backup_values = pd.to_numeric(
                                    backup_values, errors="coerce"
                                )
                        except:
                            pass

                        filled_df_indexed.loc[fill_indices, main_col] = backup_values

            # Update the main dataframe
            filled_df = filled_df_indexed.reset_index()

            # Track which cells were filled by this specific source
            after_mask = filled_df.notnull()
            filled_cells_by_source[source_name] = after_mask & ~before_mask

            after_missing = filled_df.isnull().sum().sum()
            fills_made = before_missing - after_missing
            fills_by_source[source_name] = fills_made

        except Exception as e:
            continue

    # Track which cells were filled from backups
    current_mask = filled_df.notnull()
    filled_cells = current_mask & ~original_mask

    # Convert remaining NaN and empty strings to zeros if requested
    zero_conversions = 0
    if convert_remaining_to_zero:
        for col in filled_df.columns:
            if col == index_column:
                continue

            # Handle both NaN values and empty strings
            if filled_df[col].dtype in ["object", "string"]:
                # For object/string columns, check for both NaN and empty strings
                nan_count = filled_df[col].isnull().sum()
                empty_string_count = (
                    filled_df[col].astype(str).str.strip() == ""
                ).sum()

                if nan_count > 0:
                    filled_df[col] = filled_df[col].fillna("0")
                    zero_conversions += nan_count

                # Convert empty strings to "0" for consistency
                empty_mask = filled_df[col].astype(str).str.strip() == ""
                if empty_mask.any():
                    filled_df.loc[empty_mask, col] = "0"
                    # Don't double count if they were already NaN
                    additional_empty = (
                        empty_mask.sum() - nan_count
                        if empty_mask.sum() > nan_count
                        else 0
                    )
                    zero_conversions += max(0, additional_empty)

                # Try to convert to numeric if it makes sense
                try:
                    numeric_version = pd.to_numeric(filled_df[col], errors="coerce")
                    if not numeric_version.isnull().any():
                        filled_df[col] = numeric_version
                except:
                    pass

            else:
                # For numeric columns, convert NaN to 0
                nan_count = filled_df[col].isnull().sum()
                if nan_count > 0:
                    filled_df[col] = filled_df[col].fillna(0)
                    zero_conversions += nan_count

    # Create tracking for zero-filled cells
    final_mask = filled_df.notnull()
    zero_filled_cells = (
        final_mask & ~current_mask
        if convert_remaining_to_zero
        else pd.DataFrame(False, index=filled_df.index, columns=filled_df.columns)
    )

    # Create Excel with formatting and legend
    excel_path = output_dir / output_filename
    create_formatted_excel_with_legend(
        filled_df, filled_cells_by_source, zero_filled_cells, excel_path, color_map
    )

    total_fills = sum(fills_by_source.values())
    final_missing = filled_df.isnull().sum().sum()

    print(f"Successfully filled: {total_fills:,}")
    if convert_remaining_to_zero:
        print(f"Converted to 0: {zero_conversions:,}")

    return {
        "excel_path": str(excel_path),
        "initial_missing": initial_missing,
        "total_fills": total_fills,
        "zero_conversions": zero_conversions,
        "final_missing": final_missing,
        "fills_by_source": fills_by_source,
        "color_mapping": color_map,
        "legend_created": True,
    }


def create_formatted_excel_with_legend(
    filled_df, filled_cells_by_source, zero_filled_cells, excel_path, color_map
):
    """
    Create Excel with color-coded formatting for filled cells by source and a separate Legend tab.

    Args:
        filled_df (pd.DataFrame): The main data with filled values
        filled_cells_by_source (Dict[str, pd.DataFrame]): Boolean masks for cells filled by each source
        zero_filled_cells (pd.DataFrame): Boolean mask for cells converted to zero
        excel_path (Path): Output path for Excel file
        color_map (Dict[str, str]): Mapping from source name to hex color code
    """
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    import numpy as np
    import warnings

    # Suppress openpyxl warnings
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

        # Final cleanup before Excel export
        df_clean = filled_df.copy()
        for col in df_clean.columns:
            if df_clean[col].dtype == "object":
                # Convert any remaining NaN to empty string, but keep "0" as "0"
                df_clean[col] = df_clean[col].where(pd.notna(df_clean[col]), "")
            else:
                # For numeric columns, ensure no NaN remains
                df_clean[col] = df_clean[col].where(pd.notna(df_clean[col]), 0)

        wb = Workbook()

        # === DATA TAB ===
        ws_data = wb.active
        ws_data.title = "Data"

        # Add data to worksheet
        for row in dataframe_to_rows(df_clean, index=False, header=True):
            ws_data.append(row)

        # Format filled cells by source with different colors
        for source_name, cells_mask in filled_cells_by_source.items():
            if source_name in color_map:
                color = color_map[source_name]
                font = Font(bold=True, color=color)

                # Apply formatting where cells_mask is True
                for row_idx in range(2, len(df_clean) + 2):
                    for col_idx in range(1, len(df_clean.columns) + 1):
                        if cells_mask.iloc[row_idx - 2, col_idx - 1]:
                            ws_data.cell(row=row_idx, column=col_idx).font = font

        # Format zero-converted cells with italic gray
        italic_font = Font(italic=True, color="808080")
        for row_idx in range(2, len(df_clean) + 2):
            for col_idx in range(1, len(df_clean.columns) + 1):
                if zero_filled_cells.iloc[row_idx - 2, col_idx - 1]:
                    ws_data.cell(row=row_idx, column=col_idx).font = italic_font

        # === LEGEND TAB ===
        ws_legend = wb.create_sheet("Legend")

        # Headers
        ws_legend["A1"] = "Data Source"
        ws_legend["B1"] = "Color Code"
        ws_legend["C1"] = "Description"

        # Header formatting
        header_font = Font(bold=True)
        for col in ["A1", "B1", "C1"]:
            ws_legend[col].font = header_font

        # Add color mapping to legend
        row_idx = 2
        for source_name, color_hex in color_map.items():
            # Source name
            ws_legend[f"A{row_idx}"] = source_name

            # Color code (both text color and background)
            color_cell = ws_legend[f"B{row_idx}"]
            color_cell.value = f"#{color_hex}"
            color_cell.font = Font(color=color_hex, bold=True)
            color_cell.fill = PatternFill(
                start_color=color_hex, end_color=color_hex, fill_type="solid"
            )

            # Description
            ws_legend[f"C{row_idx}"] = f"Data filled from {source_name}"

            row_idx += 1

        # Add zero-conversion entry
        ws_legend[f"A{row_idx}"] = "Zero Conversion"
        zero_cell = ws_legend[f"B{row_idx}"]
        zero_cell.value = "Italic Gray"
        zero_cell.font = Font(italic=True, color="808080")
        ws_legend[f"C{row_idx}"] = "Missing values converted to 0"

        # Auto-adjust column widths in legend
        for column in ws_legend.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws_legend.column_dimensions[column_letter].width = max_length + 2

        wb.save(excel_path)


def create_formatted_excel(filled_df, filled_cells, zero_filled_cells, excel_path):
    """
    Legacy function - Create Excel with bold formatting for filled cells and italic formatting for zero-converted cells.
    This function is kept for backward compatibility.
    """
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font
    import numpy as np
    import warnings

    # Suppress openpyxl warnings
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

        # Final cleanup before Excel export
        df_clean = filled_df.copy()
        for col in df_clean.columns:
            if df_clean[col].dtype == "object":
                # Convert any remaining NaN to empty string, but keep "0" as "0"
                df_clean[col] = df_clean[col].where(pd.notna(df_clean[col]), "")
            else:
                # For numeric columns, ensure no NaN remains
                df_clean[col] = df_clean[col].where(pd.notna(df_clean[col]), 0)

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        for row in dataframe_to_rows(df_clean, index=False, header=True):
            ws.append(row)

        # Format cells
        bold_font = Font(bold=True, color="0000FF")  # Blue and bold for backfilled
        italic_font = Font(
            italic=True, color="808080"
        )  # Gray and italic for zero-converted

        for row_idx in range(2, len(df_clean) + 2):
            for col_idx in range(1, len(df_clean.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)

                if filled_cells.iloc[row_idx - 2, col_idx - 1]:
                    cell.font = bold_font
                elif zero_filled_cells.iloc[row_idx - 2, col_idx - 1]:
                    cell.font = italic_font

        wb.save(excel_path)


def fill_missing_data(
    main_csv_path: Union[str, Path],
    backup_csv_paths: List[Union[str, Path]],
    output_path: Union[str, Path],
    output_filename: str = "filled_data.xlsx",
) -> Dict[str, Any]:
    """
    Fill missing values and create Excel file with bold formatting for filled data.

    Args:
        main_csv_path: Path to main CSV (with missing data)
        backup_csv_paths: List of backup CSV paths in priority order
        output_path: Directory to save Excel file
        output_filename: Name for Excel file
    """
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    output_dir = Path(output_path)
    output_dir.mkdir(parents=True, exist_ok=True)

    filled_df = pd.read_csv(main_csv_path)
    original_mask = filled_df.notnull()

    initial_missing = filled_df.isnull().sum().sum()

    print(f"Initial missing: {initial_missing:,}")

    fills_by_source = {}

    # Fill from each backup
    for backup_path in [Path(p) for p in backup_csv_paths]:
        backup_df = pd.read_csv(backup_path).set_index("geo_id")
        before_missing = filled_df.isnull().sum().sum()

        filled_df.set_index("geo_id", inplace=True)
        filled_df.update(backup_df, overwrite=False)  # Only fill NaN values
        filled_df.reset_index(inplace=True)

        after_missing = filled_df.isnull().sum().sum()
        fills_made = before_missing - after_missing
        fills_by_source[backup_path.stem] = fills_made

        print(f"{backup_path.stem}: {fills_made:,} fills")

    # Track which cells were filled
    current_mask = filled_df.notnull()
    filled_cells = current_mask & ~original_mask

    # Create empty zero_filled_cells DataFrame since this function doesn't convert to zeros
    zero_filled_cells = pd.DataFrame(
        False, index=filled_df.index, columns=filled_df.columns
    )

    # Create Excel with the updated function signature
    excel_path = output_dir / output_filename
    create_formatted_excel(filled_df, filled_cells, zero_filled_cells, excel_path)

    total_fills = sum(fills_by_source.values())
    final_missing = filled_df.isnull().sum().sum()

    print(f"Total: {total_fills:,} fills, {final_missing:,} remaining")

    return {
        "excel_path": str(excel_path),
        "total_fills": total_fills,
        "missing_before": initial_missing,
        "missing_after": final_missing,
        "fills_by_source": fills_by_source,
    }


def parse_query_dump_to_dataframe(file_path):
    """
    Alternative parsing method using manual token parsing.
    More robust for complex comma-separated formats.
    """
    import pandas as pd

    with open(file_path, "r") as f:
        content = f.read()

    # Remove outer brackets and clean
    content = content.strip()
    if content.startswith("[") and content.endswith("]"):
        content = content[1:-1]

    # Split into tokens, but be smart about it
    # First, let's replace newlines and extra spaces
    content = " ".join(content.split())

    queries = []
    values = []

    # Split by commas first
    parts = content.split(",")

    i = 0
    while i < len(parts):
        part = parts[i].strip()

        # If this part starts with ':', it's a query
        if part.startswith(":"):
            query = part[1:].strip()  # Remove the ':'

            # The next parts form the value until we hit another ':'
            value_parts = []
            i += 1

            # Collect value parts until we find the next query or end
            while i < len(parts):
                next_part = parts[i].strip()

                # If we hit another query, break
                if next_part.startswith(":"):
                    break

                value_parts.append(next_part)
                i += 1

            # Reconstruct the value
            if value_parts:
                # Join the parts back with commas to reconstruct the original number
                value_str = ",".join(value_parts)

                # Clean and convert
                value_cleaned = value_str.replace(",", "")
                try:
                    if "." in value_cleaned:
                        value = float(value_cleaned)
                    else:
                        value = int(value_cleaned)
                except ValueError:
                    value = value_str

                queries.append(query)
                values.append(value)

            # Don't increment i here since we already positioned it correctly
            continue
        else:
            i += 1

    return pd.DataFrame({"query": queries, "value": values})
